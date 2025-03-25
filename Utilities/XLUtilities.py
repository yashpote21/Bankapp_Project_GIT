import openpyxl


def RowCount(file, sheet_name):
    book = openpyxl.load_workbook(file)
    sheet = book[sheet_name]
    return sheet.max_row

def ColumnCount(file, sheet_name):
    book = openpyxl.load_workbook(file)
    sheet = book[sheet_name]
    return sheet.max_column

def ReadData(file, sheet_name, rownum, columnno):
    book = openpyxl.load_workbook(file)
    sheet = book[sheet_name]
    return sheet.cell(row=rownum, column=columnno).value

def WriteData(file, sheet_name, rownum, colnum, data):
    book = openpyxl.load_workbook(file)
    sheet = book[sheet_name]
    sheet.cell(row=rownum, column=colnum).value = data
    book.save(file)
